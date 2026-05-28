import streamlit as st
import chromadb
from chromadb.utils import embedding_functions
import dashscope
from dashscope import Generation
import os
import json
from io import BytesIO
import httpx
from openai import OpenAI
import pandas as pd
import re
import subprocess
import shutil
import wave
from openpyxl.utils import get_column_letter
import zipfile

# ==================== 1. 配置与初始化 ====================

# 设置页面配置
st.set_page_config(
    page_title="AI 语音语料智能处理平台",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 默认的千问 API Key
DEFAULT_API_KEY = "sk-e10cbbca48ea4f23a590884e59b3d7c9"


# ==================== 2. ChromaDB 初始化函数（提前定义）====================

@st.cache_resource
def init_chromadb():
    """初始化 ChromaDB 本地持久化客户端"""
    client = chromadb.PersistentClient(path="../ZZ_NISSAN_Script/project_own_use/Script/vdb_storage")

    # 初始化多语言嵌入模型
    ef = embedding_functions.SentenceTransformerEmbeddingFunction(
        model_name="paraphrase-multilingual-MiniLM-L12-v2"
    )

    # 获取或创建集合
    collection = client.get_or_create_collection(
        name="car_voice_kb",
        embedding_function=ef
    )

    return collection


# ==================== 3. 侧边栏 - API Key 配置 ====================

# 先定义工具函数，以便在侧边栏中使用
def get_collection_stats():
    """获取向量库统计信息"""
    try:
        # 尝试初始化 collection
        current_collection = init_chromadb()

        if not current_collection:
            return {'count': 0, 'languages': {}}

        results = current_collection.get()
        if not results or not results['ids']:
            return {'count': 0, 'languages': {}}

        import pandas as pd

        # 处理 metadatas 可能为空的情况
        if not results['metadatas'] or len(results['metadatas']) == 0:
            return {
                'count': len(results['ids']),
                'languages': {}
            }

        df_meta = pd.DataFrame(results['metadatas'])

        # 安全检查 language 列是否存在
        if 'language' not in df_meta.columns:
            return {
                'count': len(results['ids']),
                'languages': {'未知': len(results['ids'])}
            }

        lang_counts = df_meta['language'].value_counts().to_dict()

        return {
            'count': len(results['ids']),
            'languages': lang_counts
        }
    except Exception as e:
        st.error(f"获取向量库统计失败：{str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return {'count': 0, 'languages': {}}


with st.sidebar:
    st.header("⚙️ 系统配置")

    # API Key 使用模式选择
    api_key_mode = st.radio(
        "🔑 API Key 模式",
        options=["默认 API Key", "自定义 API Key"],
        help="选择使用默认配置的 API Key 或自己输入",
        index=0  # 默认选中"默认 API Key"
    )

    # 根据模式设置 API Key
    if api_key_mode == "默认 API Key":
        api_key = DEFAULT_API_KEY
        st.info("✅ 使用默认 API Key")
    else:
        api_key = st.text_input(
            "请输入您的千问 API Key",
            type="password",
            value=st.session_state.get('api_key', ''),
            help="请输入您的阿里云 DashScope API Key"
        )
        if api_key:
            st.session_state['api_key'] = api_key

    # 配置 http_client（用于 OpenAI SDK 访问）
    os.environ['NO_PROXY'] = '*'
    http_client = httpx.Client(verify=False, timeout=60.0, trust_env=False)

    # 创建 OpenAI 客户端（使用兼容模式访问千问）
    client_qwen = OpenAI(
        api_key=api_key,
        base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
        http_client=http_client
    )

    # 设置 dashscope 的 API Key（备用）
    if api_key:
        dashscope.api_key = api_key
        st.success("✅ API Key 已配置")
    else:
        st.warning("⚠️ 请先配置 API Key")

    st.divider()

    # 向量库状态
    st.header("📊 向量库状态")

    # 添加刷新按钮
    col_refresh1, col_refresh2 = st.columns([3, 1])
    with col_refresh1:
        st.write("实时统计")
    with col_refresh2:
        if st.button("🔄", key="refresh_stats_btn", help="刷新统计"):
            st.rerun()

    try:
        collection_stats = get_collection_stats()

        if collection_stats.get('count', 0) > 0:
            st.metric("语料总量", collection_stats.get('count', 0))

            if collection_stats.get('languages'):
                st.markdown("**语言分布:**")
                for lang, count in collection_stats['languages'].items():
                    st.write(f"• **{lang}**: {count} 条")

                # 尝试获取项目分布
                try:
                    current_collection = init_chromadb()
                    results = current_collection.get()

                    if results and results['metadatas']:
                        import pandas as pd

                        df_meta = pd.DataFrame(results['metadatas'])

                        if 'project' in df_meta.columns:
                            project_counts = df_meta['project'].value_counts().to_dict()
                            st.markdown("**项目分布:**")
                            for proj, count in project_counts.items():
                                st.write(f"• 📁 **{proj}**: {count} 条")
                except Exception as e:
                    pass  # 静默失败，不影响主要显示
        else:
            st.info("ℹ️ 向量库为空，请先上传语料")

    except Exception as e:
        st.error(f"❌ 读取统计失败：{str(e)}")
        st.info("💡 请尝试刷新页面或重新初始化向量库")

    st.divider()

    # 清空向量库功能
    st.header("⚠️ 数据管理")
    if st.button("🗑️ 清空所有向量库数据", key="clear_vector_db", help="删除向量库中的所有数据，此操作不可恢复！"):
        try:
            confirm_text = st.text_input(
                "请输入 'CLEAR' 确认清空所有数据",
                key="confirm_clear"
            )
            if confirm_text == 'CLEAR':
                with st.spinner("正在清空向量库..."):
                    current_collection = init_chromadb()
                    # 删除集合并重新创建
                    client = chromadb.PersistentClient(path="../ZZ_NISSAN_Script/project_own_use/Script/vdb_storage")
                    client.delete_collection(name="car_voice_kb")
                    # 重新创建空集合
                    ef = embedding_functions.SentenceTransformerEmbeddingFunction(
                        model_name="paraphrase-multilingual-MiniLM-L12-v2"
                    )
                    new_collection = client.create_collection(
                        name="car_voice_kb",
                        embedding_function=ef
                    )
                    st.success("✅ 向量库已清空！")
                    st.balloons()
                    time.sleep(1)
                    st.rerun()
            elif confirm_text:
                st.warning("⚠️ 请输入 'CLEAR'（区分大小写）来确认操作")
        except Exception as e:
            st.error(f"❌ 清空失败：{str(e)}")

    st.divider()
    st.markdown("""
    ### 📖 使用说明
    1. **用例上传**: 上传 Excel 语料文件到向量库
    2. **检索与设计**: 语义检索 + AI 生成新语料
    3. **批量标注**: 批量自动标注意图和槽位
    4. **报告分析**: 数据统计和可视化分析
    """)

# ==================== 4. 初始化集合 ====================

# 初始化集合
try:
    collection = init_chromadb()
except Exception as e:
    st.error(f"向量库初始化失败：{str(e)}")
    collection = None


# ==================== 5. 工具函数 ====================
def init_session_state():
    """初始化 session state，用于存储项目和语言选项"""
    if 'projects' not in st.session_state:
        st.session_state['projects'] = ['ZZ_NISSAN', 'DF_NISSAN', 'BYD', 'MaZda']
    if 'languages' not in st.session_state:
        st.session_state['languages'] = ['英语', '西班牙语', '阿拉伯语', '葡萄牙语', '瑞典语', '荷兰语', '意大利语',
                                         '法语', '挪威语', '波兰语']


# 初始化 session state
init_session_state()


def call_qwen_llm(system_prompt: str, user_prompt: str, model: str = "qwen-max"):
    """
    调用千问大模型（使用 OpenAI SDK）

    Args:
        system_prompt: 系统提示词
        user_prompt: 用户提示词
        model: 模型名称，默认 qwen-max

    Returns:
        str: 模型返回的内容
    """
    if not api_key:
        return "❌ 错误：请先在侧边栏配置 API Key"

    try:
        # 构造消息
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]

        # 调用千问 API（使用 OpenAI SDK）
        response = client_qwen.chat.completions.create(
            model=model,
            messages=messages
        )

        # 返回 AI 生成的内容
        return response.choices[0].message.content

    except Exception as e:
        return f"❌ 异常：{str(e)}"


def generate_unique_id(lang: str, text: str, index: int = None) -> str:
    """生成唯一 ID 用于去重"""
    import hashlib
    import time

    # 基础组合：语言 + 文本
    combined = f"{str(lang).strip().lower()}_{str(text).strip().lower()}"

    # 添加时间戳和索引以确保唯一性
    if index is not None:
        combined += f"_{index}_{time.time()}"

    return hashlib.md5(combined.encode('utf-8')).hexdigest()


# ==================== 6. 主页面布局 ====================

st.title("🚗 AI 语音语料智能处理平台")
st.markdown("---")

# 创建四个功能模块标签页
tab_upload, tab_search_design, tab_annotate, tab_analysis, tab_tools = st.tabs([
    "📤 用例上传",
    "🔍 检索与设计",
    "🪄 批量标注",
    "📈 报告分析",
    "🛠️ 工具箱"
])

# ==================== Tab 1: 用例上传 ====================

with tab_upload:
    st.header("📤 用例上传")
    st.info("上传 Excel 格式的语料文件到向量数据库，支持多语种编码")

    # 1. 项目和语言选择
    st.subheader("1️⃣ 基础信息配置")
    col1, col2 = st.columns(2)

    with col1:
        # 项目名称选择
        project_options = st.session_state.get('projects', ['默认项目'])
        selected_project = st.selectbox(
            "📁 项目名称",
            options=project_options,
            index=0 if '默认项目' in project_options else 0,
            help="选择或新建项目名称"
        )

        # 新增项目
        new_project = st.text_input(
            "➕ 新建项目",
            placeholder="输入新项目名称后点击添加",
            key="new_project_input"
        )
        if st.button("➕ 添加项目", key="add_project_btn"):
            if new_project.strip():
                if new_project not in st.session_state['projects']:
                    st.session_state['projects'].append(new_project.strip())
                    st.success(f"✅ 项目 '{new_project}' 已添加到选项")
                    st.rerun()
                else:
                    st.warning(f"⚠️ 项目 '{new_project}' 已存在")
            else:
                st.warning("⚠️ 请输入项目名称")

    with col2:
        # 语言选择
        lang_options = st.session_state.get('languages', ['zh', 'en'])
        selected_language = st.selectbox(
            "🌐 语言",
            options=lang_options,
            index=0 if 'zh' in lang_options else 0,
            help="选择或新建语言代码"
        )

        # 新增语言
        new_lang = st.text_input(
            "➕ 新建语言",
            placeholder="输入新语言代码 (如 zh, en)",
            key="new_lang_input"
        )
        if st.button("➕ 添加语言", key="add_lang_btn"):
            if new_lang.strip():
                if new_lang not in st.session_state['languages']:
                    st.session_state['languages'].append(new_lang.strip())
                    st.success(f"✅ 语言 '{new_lang}' 已添加到选项")
                    st.rerun()
                else:
                    st.warning(f"⚠️ 语言 '{new_lang}' 已存在")
            else:
                st.warning("⚠️ 请输入语言代码")

    st.divider()

    # 2. 上传 Excel 文件
    st.subheader("2️⃣ 上传 Excel 文件")
    uploaded_file = st.file_uploader(
        "选择 Excel 文件 (.xlsx)",
        type=["xlsx"],
        help="Excel 文件应包含语料数据",
        key="upload_file"
    )

    if uploaded_file:
        try:
            import pandas as pd

            # 读取 Excel 预览
            df_preview = pd.read_excel(uploaded_file)
            with st.expander("📊 查看文件内容", expanded=False):
                st.dataframe(df_preview.head(10))
                st.write(f"✅ 成功读取 {len(df_preview)} 行数据")
                st.write(f"📋 列名：{list(df_preview.columns)}")

            st.divider()

            # 3. 列名映射
            st.subheader("3️⃣ 列名映射配置")
            st.info("请指定 Excel 中各字段对应的列名")

            available_columns = list(df_preview.columns)

            col_map_1, col_map_2 = st.columns(2)

            with col_map_1:
                # Utterance 列
                utterance_col = st.selectbox(
                    "💬 Utterance (原文话术) *",
                    options=available_columns,
                    help="用户实际说的话术文本",
                    key="utterance_col"
                )

                # Translate 列
                translate_col = st.selectbox(
                    "📝 Translate (中文翻译)",
                    options=[''] + available_columns,
                    help="原文的中文翻译（可选）",
                    key="translate_col"
                )

            with col_map_2:
                # Intent 列
                intent_col = st.selectbox(
                    "🎯 Intent (意图) *",
                    options=available_columns,
                    help="语义意图标签",
                    key="intent_col"
                )

                # Slot 列
                slot_col = st.selectbox(
                    "🏷️ Slot (槽位)",
                    options=[''] + available_columns,
                    help="参数槽位信息（可选）",
                    key="slot_col"
                )

            st.divider()

            # 4. 确认并上传
            st.subheader("4️⃣ 确认并上传")

            col_btn1, col_btn2 = st.columns([1, 3])
            with col_btn1:
                upload_confirm = st.button(
                    "🚀 开始上传到向量库",
                    type="primary",
                    key="upload_confirm"
                )

            if upload_confirm:
                # 验证必填列
                if not utterance_col or not intent_col:
                    st.error("❌ 请至少选择 Utterance 和 Intent 列")
                else:
                    with st.spinner("正在处理数据并上传到向量库..."):
                        progress_bar = st.progress(0)
                        status_text = st.empty()

                        # 准备数据
                        ids, docs, metas = [], [], []
                        success_count = 0
                        skip_count = 0
                        error_count = 0

                        for idx, row in df_preview.iterrows():
                            try:
                                # 提取数据
                                utterance = str(row.get(utterance_col, '')).strip()
                                intent = str(row.get(intent_col, '')).strip()

                                # 跳过空行
                                if not utterance or not intent:
                                    skip_count += 1
                                    continue

                                # 获取可选字段
                                translate = str(row.get(translate_col, '')) if translate_col else ''
                                slot = str(row.get(slot_col, '')) if slot_col else ''

                                # 生成唯一 ID（使用索引确保唯一性）
                                uid = generate_unique_id(selected_language, utterance, idx)

                                # 构建 Document 和 Metadata
                                ids.append(uid)
                                docs.append(utterance)
                                metas.append({
                                    'project': selected_project,
                                    'language': selected_language,
                                    'intent': intent,
                                    'slot': slot,
                                    'trans': translate if translate else '',
                                    'utterance_original': utterance
                                })

                                success_count += 1

                            except Exception as e:
                                error_count += 1
                                st.error(f"行 {idx} 处理失败：{str(e)}")

                            # 更新进度
                            progress = min((idx + 1) / len(df_preview), 1.0)
                            progress_bar.progress(progress)
                            status_text.text(
                                f"处理中：{idx + 1}/{len(df_preview)} | 成功：{success_count} | 跳过：{skip_count}")

                        # 上传到 ChromaDB
                        if collection and ids:
                            try:
                                collection.upsert(ids=ids, documents=docs, metadatas=metas)

                                progress_bar.empty()
                                status_text.empty()

                                st.success(f"""
                                ### ✅ 上传完成！
                                - 📦 **成功入库**: {success_count} 条
                                - ⚠️ **跳过**: {skip_count} 条 (空数据)
                                - ❌ **失败**: {error_count} 条
                                - 📁 **项目**: {selected_project}
                                - 🌐 **语言**: {selected_language}
                                """)

                                # 显示统计
                                stats_now = get_collection_stats()
                                st.metric("当前向量库总量", stats_now['count'])

                            except Exception as e:
                                st.error(f"❌ 向量库写入失败：{str(e)}")
                        elif not collection:
                            st.error("❌ 向量库未初始化")
                        else:
                            st.warning("⚠️ 没有有效数据可上传")

        except Exception as e:
            st.error(f"❌ 文件读取失败：{str(e)}")
            import traceback

            st.code(traceback.format_exc())

# ==================== Tab 2: 检索与设计 ====================

with tab_search_design:
    st.header("🔍 检索与设计")
    st.info("基于自然语言指令的智能检索与 AI 语料设计")

    # 1. 筛选条件（可选）
    st.subheader("1️⃣ 筛选条件")
    col_filter1, col_filter2 = st.columns(2)

    with col_filter1:
        filter_project = st.selectbox(
            "📁 项目范围",
            options=['全部'] + st.session_state.get('projects', []),
            help="限定检索的项目范围"
        )

    with col_filter2:
        filter_language = st.selectbox(
            "🌐 语言范围",
            options=['全部'] + st.session_state.get('languages', []),
            help="限定检索的语言范围"
        )

    st.divider()

    # 2. 自然语言指令输入
    st.subheader("2️⃣ 输入指令")

    instruction_type = st.radio(
        "选择功能类型",
        ["🔎 查找现有 case", "✨ 设计新 case"],
        horizontal=True
    )

    natural_query = st.text_area(
        "请输入自然语言指令",
        placeholder="""
查找模式示例：
- 查找 NISSAN 项目西班牙语打开车窗的 case
- 帮我找日语中调节空调温度的说法
- 搜索英语里设置导航目的地到北京的用例

设计模式示例：
- 设计 10 条西班牙语打开车窗的 case
- 生成 15 条德语查询天气的语料，包含不同城市
- 创建 20 条法语控制音乐播放的 case，要有音量调节
        """,
        height=150,
        key="natural_query_input"
    )

    st.divider()

    # 3. 执行按钮
    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 4])
    with col_btn1:
        execute_btn = st.button("🚀 执行指令", type="primary", key="execute_search")
    with col_btn2:
        clear_btn = st.button("🗑️ 清空结果", key="clear_results")

    if clear_btn:
        if 'search_results' in st.session_state:
            del st.session_state['search_results']
        if 'generated_cases' in st.session_state:
            del st.session_state['generated_cases']
        st.rerun()

    if execute_btn and natural_query:
        if not collection:
            st.error("❌ 向量库未初始化")
        else:
            with st.spinner("🤖 AI 正在处理您的指令..."):
                try:
                    # Step 1: 从自然语言指令中提取关键信息
                    extraction_prompt = f"""你是一个语音助手 NLU 专家。请从用户指令中提取关键信息。

用户指令：{natural_query}

请提取以下信息：
1. 意图类型：是"查找"还是"设计"
2. 目标项目（如果有提及，如 NISSAN, BMW 等）
3. 目标语言（如西班牙语 es, 德语 de, 法语 fr 等）
4. 核心语义（要查找或设计什么功能的 case）
5. 数量要求（如果是设计模式，需要多少条）

请以 JSON 格式返回：
{{
    "type": "search" or "design",
    "project": "项目名或 null",
    "language": "语言代码或 null",
    "semantic_intent": "核心语义描述",
    "quantity": 数字（仅设计模式需要，默认 10）
}}
"""

                    extraction_result = call_qwen_llm(
                        system_prompt="你是一个专业的信息提取助手，只返回 JSON 数据。",
                        user_prompt=extraction_prompt,
                        model="qwen-max"
                    )

                    # 解析提取结果
                    import json

                    try:
                        extracted_info = json.loads(extraction_result)
                    except:
                        st.error(f"❌ 指令解析失败：{extraction_result}")
                        extracted_info = {
                            'type': 'search',
                            'project': None,
                            'language': None,
                            'semantic_intent': natural_query,
                            'quantity': 10
                        }

                    # 使用用户选择的筛选条件作为默认值
                    target_project = extracted_info.get('project')
                    target_language = extracted_info.get('language')

                    # 如果指令中没有指定，使用筛选器的值
                    if not target_project or target_project == '全部':
                        target_project = filter_project if filter_project != '全部' else None
                    if not target_language or target_language == '全部':
                        target_language = filter_language if filter_language != '全部' else None

                    st.info(f"""
                    📋 **指令解析结果**:
                    - 类型：{'🔍 查找' if extracted_info.get('type') == 'search' else '✨ 设计'}
                    - 项目：{target_project or '全部'}
                    - 语言：{target_language or '全部'}
                    - 语义：{extracted_info.get('semantic_intent', '')}
                    - 数量：{extracted_info.get('quantity', 10)}条
                    """)

                    # Step 2: 构建 ChromaDB 过滤条件
                    where_filter = {}
                    if target_project:
                        where_filter['project'] = target_project
                    if target_language:
                        where_filter['language'] = target_language

                    # Step 3: 向量检索
                    search_results = collection.query(
                        query_texts=[extracted_info.get('semantic_intent', natural_query)],
                        n_results=10,
                        where=where_filter if where_filter else None,
                        include=['documents', 'metadatas', 'distances']
                    )

                    # Step 4: 准备 Context
                    context_cases = []
                    if search_results['ids'] and len(search_results['ids'][0]) > 0:
                        for i in range(len(search_results['ids'][0])):
                            doc = search_results['documents'][0][i]
                            meta = search_results['metadatas'][0][i]
                            dist = search_results['distances'][0][i] if 'distances' in search_results else 0
                            score = round(1 - dist, 4)

                            context_cases.append({
                                'utterance': doc,
                                'project': meta.get('project', 'N/A'),
                                'language': meta.get('language', 'N/A'),
                                'intent': meta.get('intent', 'N/A'),
                                'slot': meta.get('slot', ''),
                                'trans': meta.get('trans', ''),
                                'similarity': score
                            })

                    # Step 5: 根据类型处理
                    if extracted_info.get('type') == 'search':
                        # === 查找模式：让大模型整理输出 ===
                        context_text = "\n".join([
                            f"{idx + 1}. [{case['language']}] {case['utterance']} | Intent: {case['intent']} | Slot: {case['slot']} | Project: {case['project']} (相似度：{case['similarity']})"
                            for idx, case in enumerate(context_cases)
                        ])

                        organize_prompt = f"""你是一个车载语音数据专家。请整理并格式化输出检索到的语料。

检索到的参考 case（共{len(context_cases)}条）：
{context_text}

请用清晰、美观的格式整理这些案例，包含：
- 项目 (Project)
- 语言 (Language)
- 原文话术 (Utterance)
- 中文翻译 (Translation)
- 意图 (Intent)
- 槽位 (Slot)
- 相似度 (Similarity Score)

以 Markdown 表格形式输出。
"""

                        organized_result = call_qwen_llm(
                            system_prompt="你是一个专业的数据整理助手，输出格式清晰美观。",
                            user_prompt=organize_prompt,
                            model="qwen-max"
                        )

                        # 保存结果到 session_state
                        st.session_state['search_results'] = context_cases

                        # 显示结果
                        st.success(f"✅ 找到 {len(context_cases)} 条相关 case！")
                        st.markdown(organized_result)

                        # 用表格展示详细数据
                        st.subheader("📊 详细数据")
                        df_results = pd.DataFrame(context_cases)
                        st.dataframe(df_results, use_container_width=True)

                    else:
                        # === 设计模式：让大模型模仿生成（增强版）===
                        quantity = extracted_info.get('quantity', 10)

                        # 增加检索数量以获取更多参考
                        extended_search = collection.query(
                            query_texts=[extracted_info.get('semantic_intent', natural_query)],
                            n_results=min(quantity * 2, 20),  # 获取更多参考
                            where=where_filter if where_filter else None,
                            include=['documents', 'metadatas', 'distances']
                        )

                        # 准备更丰富的参考案例
                        extended_cases = []
                        if extended_search['ids'] and len(extended_search['ids'][0]) > 0:
                            for i in range(len(extended_search['ids'][0])):
                                doc = extended_search['documents'][0][i]
                                meta = extended_search['metadatas'][0][i]
                                dist = extended_search['distances'][0][i] if 'distances' in extended_search else 0
                                score = round(1 - dist, 4)

                                extended_cases.append({
                                    'utterance': doc,
                                    'language': meta.get('language', 'N/A'),
                                    'intent': meta.get('intent', 'N/A'),
                                    'slot': meta.get('slot', ''),
                                    'similarity': score
                                })

                        context_text = "\n".join([
                            f"{idx + 1}. [{case['language']}] {case['utterance']} | Intent: {case['intent']} | Slot: {case['slot']} (相似度：{case['similarity']})"
                            for idx, case in enumerate(extended_cases[:min(quantity, 10)])
                        ])

                        # # 分析现有数据的标注风格
                        # intent_patterns = {}
                        # slot_patterns = {}
                        # for case in extended_cases:
                        #     intent = case['intent']
                        #     slot = case['slot']
                        #     intent_patterns[intent] = intent_patterns.get(intent, 0) + 1
                        #     if slot:
                        #         # 提取 slot key
                        #         if '=' in slot:
                        #             key = slot.split('=')[0]
                        #             slot_patterns[key] = slot_patterns.get(key, 0) + 1
                        #
                        # style_guide = ""
                        # if intent_patterns:
                        #     top_intents = sorted(intent_patterns.items(), key=lambda x: x[1], reverse=True)[:5]
                        #     style_guide += f"\n常用 Intent 风格：{', '.join([f'{i}({c}次)' for i, c in top_intents])}"
                        # if slot_patterns:
                        #     top_slots = sorted(slot_patterns.items(), key=lambda x: x[1], reverse=True)[:5]
                        #     style_guide += f"\n常用 Slot 类型：{', '.join([f'{s}({c}次)' for s, c in top_slots])}"

                        design_prompt = f"""你是一个车载语音语料设计专家。请**严格基于**向量数据库中的现有案例生成新语料。

【向量数据库参考案例】（**必须从中选择 Intent 和组合 Slot，不能自己创建！**）:
{context_text}

【**强制规则**】:
1. **Utterance 设计**:
   - ✅ 生成自然、多样化的车载场景话术
   - ✅ 保持语言风格与向量数据库中case一致
   - ❌ 避免与现有 case 完全重复
   - ✅ 可以改变句式、用词，但语义要与主题相关，且是连续的句子，不要含有逗号之类的符号

2. **Intent 标注（强制执行）**:
   - ✅ **必须**从参考案例中选择一个现有的 Intent
   - ❌ **禁止**创建新的 Intent 名称
   - ✅ 如果语义接近某个参考案例，直接使用其 Intent
   - ✅ 如果有多个可能的 Intent，选择出现频率最高的那个

3. **Slot 标注（强制执行）**:
   - ✅ 格式：key=value（key 必须是参考案例中出现过的），多个slot用分号隔开
   - ✅ 如果没有合适的参数，保持为空字符串 ""


5. **生成要求**:
   - 生成{quantity}条新的语料
   - 主题：{extracted_info.get('semantic_intent', '相关功能')}
   - 目标语言：{target_language or '自动判断'}
请以 JSON 数组格式返回：
[
  {{
    "utterance": "打开主驾驶车窗",
    "trans": "打开主驾驶侧的车窗",
    "intent": "settings_and_control:open:windows",
    "slot": "area=driver",
    "language": "中文",
    "project": "NISSAN"
  }}
]
"""

                        generated_result = call_qwen_llm(
                            system_prompt="你是一个专业的车载语音语料生成专家，严格遵循现有数据的标注规范，只返回 JSON 数组。",
                            user_prompt=design_prompt,
                            model="qwen-max"
                        )

                        # 解析生成的结果
                        try:
                            generated_cases = json.loads(generated_result)
                            if isinstance(generated_cases, dict) and 'items' in generated_cases:
                                generated_cases = generated_cases['items']

                            # # === 后处理验证：确保 Intent 和 Slot 来自参考案例 ===
                            # if extended_cases:
                            #     # 收集所有参考案例中的 Intent 和 Slot keys
                            #     valid_intents = set([case['intent'] for case in extended_cases])
                            #     valid_slot_keys = set()
                            #     for case in extended_cases:
                            #         if case['slot']:
                            #             keys = [s.split('=')[0].strip() for s in case['slot'].split(',')]
                            #             valid_slot_keys.update(keys)
                            #
                            #     # 逐条验证生成的 case
                            #     for i, case in enumerate(generated_cases):
                            #         # 验证 Intent
                            #         if 'intent' in case and case['intent'] not in valid_intents:
                            #             # 使用最接近的 Intent（基于语义相似度）
                            #             best_intent = extended_cases[0]['intent']
                            #             case['intent'] = best_intent
                            #             if '_validation_notes' not in case:
                            #                 case['_validation_notes'] = ''
                            #             case[
                            #                 '_validation_notes'] += f"Intent 已修正为 '{best_intent}'（原值不在参考案例中）; "
                            #
                            #         # 验证 Slot
                            #         if 'slot' in case and case['slot']:
                            #             original_slot = case['slot']
                            #             new_slots = []
                            #
                            #             for slot_pair in original_slot.split(','):
                            #                 if '=' in slot_pair:
                            #                     key, value = slot_pair.split('=', 1)
                            #                     key = key.strip()
                            #                     value = value.strip()
                            #
                            #                     if key not in valid_slot_keys:
                            #                         if '_validation_notes' not in case:
                            #                             case['_validation_notes'] = ''
                            #                         case['_validation_notes'] += f"过滤无效 slot key '{key}'; "
                            #                     else:
                            #                         new_slots.append(f"{key}={value}")
                            #                 else:
                            #                     if slot_pair.strip() in valid_slot_keys:
                            #                         new_slots.append(slot_pair.strip())
                            #
                            #             case['slot'] = ','.join(new_slots)

                        except Exception as e:
                            st.error(f"❌ 结果解析失败：{generated_result}")
                            generated_cases = []

                        # 保存到 session_state
                        st.session_state['generated_cases'] = generated_cases

                        # 显示结果
                        st.success(f"✅ AI 生成了 {len(generated_cases)} 条新语料！")

                        # 使用可编辑表格展示
                        if generated_cases:
                            st.subheader("📝 AI 生成结果（可编辑）")
                            st.info("💡 您可以直接点击表格修改 Intent 和 Slot 字段")

                            df_generated = pd.DataFrame(generated_cases)

                            # 确保必要的列存在
                            required_cols = ['utterance', 'trans', 'intent', 'slot', 'language']
                            for col in required_cols:
                                if col not in df_generated.columns:
                                    df_generated[col] = ''

                            # 添加项目列（如果有指定）
                            if target_project and 'project' not in df_generated.columns:
                                df_generated['project'] = target_project
                            elif 'project' not in df_generated.columns:
                                df_generated['project'] = ''

                            # 创建可编辑表格
                            edited_df = st.data_editor(
                                df_generated,
                                use_container_width=True,
                                num_rows="dynamic",
                                column_config={
                                    "utterance": st.column_config.TextColumn(
                                        "💬 话术 (Utterance)",
                                        help="用户实际说的话",
                                        width="large"
                                    ),
                                    "trans": st.column_config.TextColumn(
                                        "📝 中文翻译",
                                        help="中文含义"
                                    ),
                                    "intent": st.column_config.TextColumn(
                                        "🎯 意图 (Intent)",
                                        help="语义意图标签",
                                        width="medium"
                                    ),
                                    "slot": st.column_config.TextColumn(
                                        "🏷️ 槽位 (Slot)",
                                        help="参数槽位信息",
                                        width="medium"
                                    ),
                                    "language": st.column_config.TextColumn(
                                        "🌐 语言",
                                        help="语言代码",
                                        width="small"
                                    ),
                                    "project": st.column_config.TextColumn(
                                        "📁 项目",
                                        help="所属项目",
                                        width="small"
                                    )
                                },
                                hide_index=True
                            )

                            # 更新 session_state
                            st.session_state['generated_cases_edited'] = edited_df

                            # 操作按钮
                            st.divider()
                            st.subheader("💾 导出与入库")

                            col_download, col_upload = st.columns(2)

                            with col_download:
                                # 下载按钮
                                output = BytesIO()
                                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                    edited_df.to_excel(writer, index=False)

                                st.download_button(
                                    label="📥 下载 Excel",
                                    data=output.getvalue(),
                                    file_name=f"AI_Generated_Cases_{target_language or 'mixed'}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                            with col_upload:
                                # 一键入库按钮
                                if st.button("➕ 一键同步到向量库", type="primary"):
                                    with st.spinner("正在上传到向量库..."):
                                        ids, docs, metas = [], [], []

                                        for idx, row in edited_df.iterrows():
                                            utterance = str(row.get('utterance', '')).strip()
                                            if not utterance:
                                                continue

                                            # 使用索引生成唯一 ID
                                            uid = generate_unique_id(
                                                str(row.get('language', 'zh')),
                                                utterance,
                                                idx
                                            )
                                            ids.append(uid)
                                            docs.append(utterance)
                                            metas.append({
                                                'project': str(row.get('project', '')),
                                                'language': str(row.get('language', 'zh')),
                                                'intent': str(row.get('intent', '')),
                                                'slot': str(row.get('slot', '')),
                                                'trans': str(row.get('trans', '')),
                                                'utterance_original': utterance
                                            })

                                        if ids and collection:
                                            collection.upsert(ids=ids, documents=docs, metadatas=metas)
                                            st.success(f"✅ 成功上传 {len(ids)} 条语料到向量库！")
                                            st.rerun()
                                        else:
                                            st.warning("⚠️ 没有有效数据可上传")

                        else:
                            st.warning("⚠️ AI 未生成有效结果，请调整指令后重试")

                except Exception as e:
                    st.error(f"❌ 处理失败：{str(e)}")
                    import traceback

                    st.code(traceback.format_exc())

    elif execute_btn and not natural_query:
        st.warning("⚠️ 请输入指令内容")

    # 显示历史结果（如果有）
    if 'search_results' in st.session_state and execute_btn is False:
        pass  # 可以选择是否显示历史结果

    if 'generated_cases_edited' in st.session_state and execute_btn is False:
        pass  # 可以选择是否显示历史结果

# ==================== Tab 3: 批量标注 ====================

with tab_annotate:
    st.header("🪄 批量标注")
    st.info("上传待标注 Excel，AI 自动识别意图和槽位，支持结果编辑和导出")

    # 1. 上传文件
    st.subheader("1️⃣ 上传待标注文件")
    task_file = st.file_uploader(
        "选择 Excel 文件 (.xlsx)",
        type=["xlsx"],
        help="Excel 文件应包含 Utterance 列，其他列可选",
        key="annotate_upload"
    )

    if task_file:
        try:
            import pandas as pd
            from io import BytesIO

            # 读取原始 Excel（保留所有原始列）
            df_original = pd.read_excel(task_file)

            with st.expander("📊 查看文件内容", expanded=False):
                st.dataframe(df_original.head(10))
                st.write(f"✅ 成功读取 {len(df_original)} 行数据")
                st.write(f"📋 所有列名：{list(df_original.columns)}")

            st.divider()

            # 2. 列名映射
            st.subheader("2️⃣ 列名映射配置")
            st.info("请指定 Excel 中各字段对应的列名（Intent 和 Slot 可为空，将自动生成新列）")

            available_columns = list(df_original.columns)

            col_map_1, col_map_2 = st.columns(2)

            with col_map_1:
                # Utterance 列（必填）
                annotate_utterance_col = st.selectbox(
                    "💬 Utterance (原文话术) *",
                    options=available_columns,
                    help="需要标注的用户话术文本",
                    key="annotate_utterance_col"
                )

                # 语言列（可选，用于限定检索范围）
                annotate_lang_col = st.selectbox(
                    "🌐 Language (语言代码，可选)",
                    options=[''] + available_columns,
                    help="如果有语言列，将按此语言检索；否则自动判断",
                    key="annotate_lang_col"
                )

            with col_map_2:
                # Intent 列（可选，如果为空会创建新列）
                annotate_intent_col = st.selectbox(
                    "🎯 Intent (意图列，可选)",
                    options=[''] + available_columns,
                    help="如果选择现有列，将在该列填充；否则创建新列 'AI_Intent'",
                    key="annotate_intent_col"
                )

                # Slot 列（可选，如果为空会创建新列）
                annotate_slot_col = st.selectbox(
                    "🏷️ Slot (槽位列，可选)",
                    options=[''] + available_columns,
                    help="如果选择现有列，将在该列填充；否则创建新列 'AI_Slot'",
                    key="annotate_slot_col"
                )

            st.divider()

            # 3. 标注设置
            st.subheader("3️⃣ 标注设置")

            col_setting1, col_setting2 = st.columns(2)

            with col_setting1:
                batch_size = st.slider(
                    "批处理大小",
                    min_value=1,
                    max_value=50,
                    value=10,
                    step=1,
                    help="每次处理的条数，影响进度显示频率"
                )

            with col_setting2:
                top_k_results = st.slider(
                    "检索参考数量",
                    min_value=1,
                    max_value=10,
                    value=3,
                    step=1,
                    help="为每条 Utterance 检索多少条相似参考 case"
                )

            st.divider()

            # 4. 开始标注按钮
            col_btn1, col_btn2 = st.columns([1, 4])
            with col_btn1:
                start_annotate_btn = st.button(
                    "🚀 开始 AI 标注",
                    type="primary",
                    key="start_annotate"
                )

            if start_annotate_btn:
                if not annotate_utterance_col:
                    st.error("❌ 请选择 Utterance 列")
                elif not collection:
                    st.error("❌ 向量库未初始化")
                else:
                    with st.spinner("🤖 AI 正在批量标注中..."):
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        info_text = st.empty()

                        # 复制原始 DataFrame（保留所有原始列）
                        df_result = df_original.copy()

                        # 确定输出列名
                        intent_output_col = annotate_intent_col if annotate_intent_col else 'AI_Intent'
                        slot_output_col = annotate_slot_col if annotate_slot_col else 'AI_Slot'
                        confidence_col = 'AI_Confidence'
                        reason_col = 'AI_Reference'

                        # 初始化结果列
                        df_result[intent_output_col] = ''
                        df_result[slot_output_col] = ''
                        df_result[confidence_col] = 0.0
                        df_result[reason_col] = ''

                        total_rows = len(df_result)
                        success_count = 0
                        skip_count = 0
                        error_count = 0

                        # 逐行处理
                        for idx, row in df_result.iterrows():
                            try:
                                utterance = str(row.get(annotate_utterance_col, '')).strip()

                                # 跳过空行
                                if not utterance:
                                    skip_count += 1
                                    df_result.at[idx, intent_output_col] = 'EMPTY'
                                    df_result.at[idx, slot_output_col] = 'EMPTY'
                                    continue

                                # 获取语言（如果有语言列）
                                target_lang = None
                                if annotate_lang_col:
                                    target_lang = str(row.get(annotate_lang_col, '')).strip()

                                # Step 1: 向量检索最相似的 3 条已知 case
                                where_filter = None
                                if target_lang:
                                    where_filter = {'language': target_lang}

                                search_results = collection.query(
                                    query_texts=[utterance],
                                    n_results=top_k_results,
                                    where=where_filter,
                                    include=['documents', 'metadatas', 'distances']
                                )

                                # Step 2: 准备参考 Context
                                reference_cases = []
                                if search_results['ids'] and len(search_results['ids'][0]) > 0:
                                    for i in range(len(search_results['ids'][0])):
                                        doc = search_results['documents'][0][i]
                                        meta = search_results['metadatas'][0][i]
                                        dist = search_results['distances'][0][i] if 'distances' in search_results else 0
                                        score = round(1 - dist, 4)

                                        reference_cases.append({
                                            'utterance': doc,
                                            'language': meta.get('language', 'N/A'),
                                            'intent': meta.get('intent', 'N/A'),
                                            'slot': meta.get('slot', ''),
                                            'similarity': score
                                        })

                                # Step 3: 调用千问大模型判断 Intent 和 Slot（严格匹配版）
                                reference_text = "\n".join([
                                    f"{i + 1}. [{case['language']}] {case['utterance']} | Intent: {case['intent']} | Slot: {case['slot']} (相似度：{case['similarity']})"
                                    for i, case in enumerate(reference_cases)
                                ]) if reference_cases else "无参考案例"

                                annotation_prompt = f"""你是一个车载语音 NLU 标注专家。请**严格基于**向量数据库中的参考案例进行标注。

【待标注话术】:
{utterance}
{f"目标语言：{target_lang}" if target_lang else ""}

【向量数据库参考案例】（**必须从中选择 Intent 和 Slot，不能自己创建！**）:
{reference_text}

【**强制标注规则**】:
1. **Intent 选择（强制执行）**:
   - ✅ **必须**从参考案例中选择与待标注话术最相似的 Intent
   - ❌ **禁止**创建新的 Intent 名称
   - ✅ 如果参考案例有多个不同的 Intent，选择语义最接近的那个
   - ✅ 如果没有参考案例（相似度都很低），标记为 "UNKNOWN"

2. **Slot 提取（强制执行）**:
   - ✅ **必须**从参考案例中提取 Slot 格式和参数名
   - ❌ **禁止**创建新的 Slot 参数名
   - ✅ 如果话术中包含参数，使用参考案例中已有的 Slot key（如 temperature、volume 等）
   - ✅ 如果没有合适的 Slot 参考，保持为空字符串 ""

3. **Utterance 语义匹配**:
   - 分析待标注话术与哪个参考案例语义最相似
   - 在 reason 中明确指出参考了哪个案例

4. **置信度评估**:
   - 0.9-1.0: 有非常相似的参考 case，直接使用其 Intent 和 Slot
   - 0.7-0.9: 有参考价值，Intent 相同，Slot 可能需要调整
   - 0.5-0.7: 参考较少，但 Intent 应该一致
   - <0.5: 缺乏参考，标记为 UNKNOWN


请以 JSON 格式返回：
{{
    "intent": "必须从参考案例中选择的 Intent 名称",
    "slot": "使用参考案例中的 Slot格式，如 'temperature=26'",
    "confidence": 0.92,
    "reason": "参考案例 1 语义最相似（相似度 0.92），直接使用其 Intent: CLIMATE_CONTROL"
}}
"""

                                llm_result = call_qwen_llm(
                                    system_prompt="你是一个专业的语音标注专家，只返回 JSON 数据。",
                                    user_prompt=annotation_prompt,
                                    model="qwen-max"
                                )

                                # 解析 LLM 结果
                                import json

                                try:
                                    annotation_data = json.loads(llm_result)
                                    ai_intent = annotation_data.get('intent', 'UNKNOWN')
                                    ai_slot = annotation_data.get('slot', '')
                                    ai_confidence = annotation_data.get('confidence', 0.0)
                                    ai_reason = annotation_data.get('reason', '')

                                    # === 后处理验证：确保 Intent 和 Slot 来自参考案例 ===
                                    if reference_cases:
                                        # 收集所有参考案例中的 Intent 和 Slot keys
                                        valid_intents = set([case['intent'] for case in reference_cases])
                                        valid_slot_keys = set()
                                        for case in reference_cases:
                                            if case['slot']:
                                                # 提取 slot key (如 "temperature=25" -> "temperature")
                                                keys = [s.split('=')[0].strip() for s in case['slot'].split(',')]
                                                valid_slot_keys.update(keys)

                                        # 验证 Intent 是否在参考案例中
                                        if ai_intent not in valid_intents:
                                            # 如果 AI 自己创建了新的 Intent，使用最佳匹配的 Intent
                                            best_intent = reference_cases[0]['intent']
                                            ai_reason += f" | 修正：原 Intent '{ai_intent}' 不在参考案例中，已改为 '{best_intent}'"
                                            ai_intent = best_intent

                                        # 验证 Slot key 是否在参考案例中
                                        if ai_slot:
                                            new_slots = []
                                            for slot_pair in ai_slot.split(','):
                                                if '=' in slot_pair:
                                                    key, value = slot_pair.split('=', 1)
                                                    key = key.strip()
                                                    value = value.strip()

                                                    if key not in valid_slot_keys:
                                                        # 如果 key 不在参考案例中，跳过这个 slot
                                                        ai_reason += f" | 过滤：无效 slot key '{key}'"
                                                    else:
                                                        new_slots.append(f"{key}={value}")
                                                else:
                                                    # 没有值的 slot，检查是否有效
                                                    if slot_pair.strip() in valid_slot_keys:
                                                        new_slots.append(slot_pair.strip())

                                            ai_slot = ','.join(new_slots)

                                except Exception as e:
                                    # 如果解析失败，使用第一个参考案例的值
                                    if reference_cases:
                                        ai_intent = reference_cases[0]['intent']
                                        ai_slot = reference_cases[0]['slot']
                                        ai_confidence = reference_cases[0]['similarity']
                                        ai_reason = f"基于最佳匹配 (相似度：{ai_confidence})"
                                    else:
                                        ai_intent = 'NO_REFERENCE'
                                        ai_slot = ''
                                        ai_confidence = 0.0
                                        ai_reason = "向量库中无参考案例"

                                # 填充结果
                                df_result.at[idx, intent_output_col] = ai_intent
                                df_result.at[idx, slot_output_col] = ai_slot
                                df_result.at[idx, confidence_col] = round(ai_confidence, 4)
                                df_result.at[idx, reason_col] = ai_reason

                                success_count += 1

                            except Exception as e:
                                error_count += 1
                                df_result.at[idx, intent_output_col] = f'ERROR: {str(e)[:50]}'
                                df_result.at[idx, slot_output_col] = 'ERROR'

                            # 更新进度
                            progress = min((idx + 1) / total_rows, 1.0)
                            progress_bar.progress(progress)
                            status_text.text(
                                f"处理进度：{idx + 1}/{total_rows} | 成功：{success_count} | 跳过：{skip_count} | 错误：{error_count}")

                            # 每处理 batch_size 条，显示一次详细信息
                            if (idx + 1) % batch_size == 0:
                                info_text.info(
                                    f"已处理 {idx + 1}/{total_rows} 条，当前成功率：{success_count / (idx + 1) * 100:.1f}%")

                        progress_bar.empty()
                        status_text.empty()
                        info_text.empty()

                        # 保存到 session_state
                        st.session_state['annotate_result'] = df_result
                        st.session_state['annotate_intent_col'] = intent_output_col
                        st.session_state['annotate_slot_col'] = slot_output_col

                        # 显示结果
                        st.success(f"""
                        ### ✅ 标注完成！
                        - 📦 **总计**: {total_rows} 条
                        - ✅ **成功**: {success_count} 条
                        - ⚠️ **跳过**: {skip_count} 条
                        - ❌ **错误**: {error_count} 条
                        - 📊 **新增列**: `{intent_output_col}`, `{slot_output_col}`, `{confidence_col}`, `{reason_col}`
                        """)

                        # 显示可编辑表格
                        st.subheader("📝 标注结果 Review（可编辑）")
                        st.info("💡 您可以直接点击表格修改 Intent、Slot 等字段")

                        # 配置列显示样式
                        column_config = {
                            intent_output_col: st.column_config.TextColumn(
                                "🎯 意图 (Intent)",
                                help="AI 识别的意图标签",
                                width="medium"
                            ),
                            slot_output_col: st.column_config.TextColumn(
                                "🏷️ 槽位 (Slot)",
                                help="AI 识别的槽位信息",
                                width="medium"
                            ),
                            confidence_col: st.column_config.NumberColumn(
                                "📊 置信度",
                                format="%.2f",
                                min_value=0,
                                max_value=1,
                                width="small"
                            ),
                            reason_col: st.column_config.TextColumn(
                                "📝 参考说明",
                                help="AI 标注的理由",
                                width="large"
                            )
                        }

                        edited_result = st.data_editor(
                            df_result,
                            use_container_width=True,
                            num_rows="dynamic",
                            column_config=column_config,
                            hide_index=True
                        )

                        # 更新 session_state
                        st.session_state['annotate_result_edited'] = edited_result

                        # 下载按钮
                        st.divider()
                        st.subheader("💾 导出结果")

                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            edited_result.to_excel(writer, index=False)

                        st.download_button(
                            label="📥 下载标注后的 Excel",
                            data=output.getvalue(),
                            file_name=f"AI_Annotated_{task_file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )

                        st.info("💡 下载的 Excel 保留了所有原始列，并新增了 AI 标注的 Intent、Slot、置信度和参考说明列")

        except Exception as e:
            st.error(f"❌ 文件处理失败：{str(e)}")
            import traceback

            st.code(traceback.format_exc())

    elif task_file:
        st.info("👆 请点击上方 **开始标注** 按钮开始处理")

# ==================== Tab 4: 报告分析 ====================

with tab_analysis:
    st.header("📈 测试报告分析")
    st.info("上传测试报告 Excel，AI 智能分析失败原因")

    # 1. 上传测试报告
    st.subheader("1️⃣ 上传测试报告")
    report_file = st.file_uploader(
        "选择 Excel 测试报告 (.xlsx)",
        type=["xlsx"],
        help="Excel 文件应包含测试结果数据",
        key="report_upload"
    )

    if report_file:
        try:
            import pandas as pd
            from io import BytesIO

            # 读取 Excel 获取所有 sheet 名
            excel_file = pd.ExcelFile(report_file)
            all_sheets = excel_file.sheet_names

            st.divider()

            # 2. Sheet 选择和列配置
            st.subheader("2️⃣ 配置数据列")

            col_config1, col_config2 = st.columns(2)

            with col_config1:
                # Sheet 选择
                selected_sheet = st.selectbox(
                    "📊 选择工作表 (Sheet)",
                    options=all_sheets,
                    index=0,
                    help="选择包含测试数据的 sheet"
                )

                # 读取选中的 sheet
                df_report_raw = pd.read_excel(report_file, sheet_name=selected_sheet)

                with st.expander("📋 查看数据预览", expanded=False):
                    st.dataframe(df_report_raw.head(10))
                    st.write(f"✅ 成功读取 {len(df_report_raw)} 行数据")
                    st.write(f"📝 所有列名：{list(df_report_raw.columns)}")

                available_columns = list(df_report_raw.columns)

            with col_config2:
                # 失败结果列
                final_res_col = st.selectbox(
                    "❌ 最终结果列 (final_res)",
                    options=available_columns,
                    help="标识测试结果的列（F/Fail 表示失败）",
                    index=available_columns.index('final_res') if 'final_res' in available_columns else 0
                )

            st.divider()

            # 3. 列名映射
            st.subheader("3️⃣ 关键列映射")
            st.info("请指定以下 8 个关键字段对应的列名")

            col_map_1, col_map_2, col_map_3 = st.columns(3)

            with col_map_1:
                st.markdown("**ASR 相关列**")
                ref_asr_col = st.selectbox(
                    "🎤 参考 ASR (ref_asr)",
                    options=available_columns,
                    help="标准文本/参考识别结果",
                    index=available_columns.index('ref_asr') if 'ref_asr' in available_columns else 0,
                    key="ref_asr_col"
                )
                actual_asr_col = st.selectbox(
                    "🎤 实际 ASR (actual_asr)",
                    options=available_columns,
                    help="实际识别结果",
                    index=available_columns.index('actual_asr') if 'actual_asr' in available_columns else 0,
                    key="actual_asr_col"
                )

            with col_map_2:
                st.markdown("**Intent 相关列**")
                ref_intent_col = st.selectbox(
                    "🎯 参考 Intent (ref_intent)",
                    options=available_columns,
                    help="标准意图标签",
                    index=available_columns.index('ref_intent') if 'ref_intent' in available_columns else 0,
                    key="ref_intent_col"
                )
                actual_intent_col = st.selectbox(
                    "🎯 实际 Intent (actual_intent)",
                    options=available_columns,
                    help="实际识别的意图",
                    index=available_columns.index('actual_intent') if 'actual_intent' in available_columns else 0,
                    key="actual_intent_col"
                )

            with col_map_3:
                st.markdown("**其他关键列**")
                # 可选的其他列
                ref_slot_col = st.selectbox(
                    "🏷️ 参考 Slot (ref_slot，可选)",
                    options=[''] + available_columns,
                    help="标准槽位信息",
                    index=available_columns.index('ref_slot') if 'ref_slot' in available_columns else 0,
                    key="ref_slot_col"
                )
                actual_slot_col = st.selectbox(
                    "🏷️ 实际 Slot (actual_slot，可选)",
                    options=[''] + available_columns,
                    help="实际槽位信息",
                    index=available_columns.index('actual_slot') if 'actual_slot' in available_columns else 0,
                    key="actual_slot_col"
                )

            st.divider()

            # 4. 分析设置
            st.subheader("4️⃣ 分析设置")

            col_setting1, col_setting2 = st.columns(2)

            with col_setting1:
                top_k_for_analysis = st.slider(
                    "检索参考数量",
                    min_value=1,
                    max_value=10,
                    value=5,
                    step=1,
                    help="为每条失败数据检索多少条标准 case 作为参考"
                )

            with col_setting2:
                fail_filter = st.selectbox(
                    "失败标识筛选",
                    options=['F', 'Fail', 'Both'],
                    index=2,
                    help="筛选哪些标识的行作为失败数据"
                )

            st.divider()

            # 5. 开始分析按钮
            col_btn1, col_btn2 = st.columns([1, 4])
            with col_btn1:
                start_analysis_btn = st.button(
                    "🚀 开始 AI 分析",
                    type="primary",
                    key="start_report_analysis"
                )

            if start_analysis_btn:
                if not collection:
                    st.error("❌ 向量库未初始化")
                else:
                    with st.spinner("🤖 AI 正在智能分析失败原因..."):
                        progress_bar = st.progress(0)
                        status_text = st.empty()

                        # 筛选失败数据
                        fail_values = []
                        if fail_filter == 'F':
                            fail_values = ['F']
                        elif fail_filter == 'Fail':
                            fail_values = ['Fail']
                        else:
                            fail_values = ['F', 'Fail']

                        # 转换为字符串后筛选
                        df_report_raw[final_res_col] = df_report_raw[final_res_col].astype(str)
                        df_failed = df_report_raw[df_report_raw[final_res_col].isin(fail_values)].copy()

                        total_failed = len(df_failed)

                        if total_failed == 0:
                            st.warning(f"⚠️ 未找到失败数据（筛选条件：{fail_filter}")
                        else:
                            st.info(f"📊 共找到 {total_failed} 条失败数据，开始 AI 分析...")

                            analysis_results = []
                            success_count = 0
                            error_count = 0

                            for idx, row in df_failed.iterrows():
                                try:
                                    # 提取关键信息
                                    ref_asr = str(row.get(ref_asr_col, '')).strip()
                                    actual_asr = str(row.get(actual_asr_col, '')).strip()
                                    ref_intent = str(row.get(ref_intent_col, '')).strip()
                                    actual_intent = str(row.get(actual_intent_col, '')).strip()

                                    # 可选的 slot 信息
                                    ref_slot = str(row.get(ref_slot_col, '')) if ref_slot_col else ''
                                    actual_slot = str(row.get(actual_slot_col, '')) if actual_slot_col else ''

                                    # Step 1: 使用 ref_asr 或 ref_intent 在向量库中检索标准 case
                                    search_query = ref_asr if ref_asr else ref_intent

                                    search_results = collection.query(
                                        query_texts=[search_query],
                                        n_results=top_k_for_analysis,
                                        include=['documents', 'metadatas', 'distances']
                                    )

                                    # Step 2: 准备标准数据 Context
                                    standard_cases = []
                                    if search_results['ids'] and len(search_results['ids'][0]) > 0:
                                        for i in range(len(search_results['ids'][0])):
                                            doc = search_results['documents'][0][i]
                                            meta = search_results['metadatas'][0][i]
                                            dist = search_results['distances'][0][
                                                i] if 'distances' in search_results else 0
                                            score = round(1 - dist, 4)

                                            standard_cases.append({
                                                'utterance': doc,
                                                'intent': meta.get('intent', 'N/A'),
                                                'slot': meta.get('slot', ''),
                                                'language': meta.get('language', 'N/A'),
                                                'similarity': score
                                            })

                                    # Step 3: 调用千问大模型分析失败原因
                                    standard_text = "\n".join([
                                        f"{i + 1}. [{case['language']}] {case['utterance']} | Intent: {case['intent']} | Slot: {case['slot']} (相似度：{case['similarity']})"
                                        for i, case in enumerate(standard_cases)
                                    ]) if standard_cases else "无标准参考案例"

                                    analysis_prompt = f"""你是一个车载语音测试分析专家。请分析测试失败的原因。

【标准数据】（来自向量数据库）：
{standard_text}

【实际失败数据】:
- 参考 ASR: {ref_asr}
- 实际 ASR: {actual_asr}
- 参考 Intent: {ref_intent}
- 实际 Intent: {actual_intent}
{f"- 参考 Slot: {ref_slot}" if ref_slot else ""}
{f"- 实际 Slot: {actual_slot}" if actual_slot else ""}

请分析：
1. 失败的主要原因是什么？（例如：ASR 识别错误、意图跳转、模型拒识等）
2. 是 ASR 问题还是 NLU 问题？
3. 给出改进建议

请以 JSON 格式返回：
{{
    "failure_type": "ASR 错误" or "NLU 错误" or "数据标注问题" or "其他",
    "reason": "详细的失败原因分析",
    "suggestion": "改进建议",
    "confidence": 0.95
}}
"""

                                    llm_result = call_qwen_llm(
                                        system_prompt="你是一个专业的测试分析专家，只返回 JSON 数据。",
                                        user_prompt=analysis_prompt,
                                        model="qwen-max"
                                    )

                                    # 解析分析结果
                                    import json

                                    try:
                                        analysis_data = json.loads(llm_result)
                                        failure_type = analysis_data.get('failure_type', 'Unknown')
                                        reason = analysis_data.get('reason', '')
                                        suggestion = analysis_data.get('suggestion', '')
                                        confidence = analysis_data.get('confidence', 0.0)
                                    except:
                                        failure_type = 'Analysis Failed'
                                        reason = f"AI 分析失败：{llm_result[:200]}"
                                        suggestion = '请人工复核'
                                        confidence = 0.0

                                    # 保存分析结果
                                    result_row = {
                                        '序号': idx + 1,
                                        '参考 ASR': ref_asr,
                                        '实际 ASR': actual_asr,
                                        '参考 Intent': ref_intent,
                                        '实际 Intent': actual_intent,
                                        '失败类型': failure_type,
                                        '原因分析': reason,
                                        '改进建议': suggestion,
                                        '置信度': round(confidence, 2),
                                        '标准案例数': len(standard_cases)
                                    }

                                    # 添加原始数据的所有列
                                    for col in df_failed.columns:
                                        result_row[f'原_{col}'] = row[col]

                                    analysis_results.append(result_row)
                                    success_count += 1

                                except Exception as e:
                                    error_count += 1
                                    st.error(f"行 {idx} 分析失败：{str(e)}")

                                # 更新进度
                                progress = min((idx + 1) / total_failed, 1.0)
                                progress_bar.progress(progress)
                                status_text.text(
                                    f"分析进度：{idx + 1}/{total_failed} | 成功：{success_count} | 错误：{error_count}")

                            progress_bar.empty()
                            status_text.empty()

                            # 显示结果
                            st.success(f"""
                            ### ✅ 分析完成！
                            - 📊 **失败数据总数**: {total_failed}
                            - ✅ **分析成功**: {success_count} 条
                            - ❌ **分析失败**: {error_count} 条
                            """)

                            # 汇总统计
                            if analysis_results:
                                df_analysis = pd.DataFrame(analysis_results)

                                # 失败类型统计
                                st.subheader("📊 失败类型分布")
                                failure_type_dist = df_analysis['失败类型'].value_counts()

                                col_stat1, col_stat2 = st.columns(2)
                                with col_stat1:
                                    st.metric("主要失败类型",
                                              failure_type_dist.index[0] if len(failure_type_dist) > 0 else 'N/A')
                                with col_stat2:
                                    st.metric("占比",
                                              f"{failure_type_dist.iloc[0] / len(df_analysis) * 100:.1f}%" if len(
                                                  failure_type_dist) > 0 else 'N/A')

                                # 可视化
                                st.bar_chart(failure_type_dist)

                                st.divider()

                                # 详细结果表格
                                st.subheader("📝 详细分析结果")

                                # 只显示关键列
                                display_cols = ['序号', '参考 ASR', '实际 ASR', '参考 Intent', '实际 Intent',
                                                '失败类型', '原因分析', '改进建议', '置信度']
                                available_display_cols = [col for col in display_cols if col in df_analysis.columns]

                                st.dataframe(
                                    df_analysis[available_display_cols],
                                    use_container_width=True,
                                    hide_index=True
                                )

                                # 保存到 session_state
                                st.session_state['report_analysis_result'] = df_analysis

                                st.divider()

                                # 导出按钮
                                st.subheader("💾 导出分析报告")

                                output = BytesIO()
                                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                    df_analysis.to_excel(writer, index=False)

                                st.download_button(
                                    label="📥 下载分析报告 (Excel)",
                                    data=output.getvalue(),
                                    file_name=f"AI_Analysis_Report_{report_file.name}",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )

                                st.info("💡 导出的 Excel 包含所有原始列和 AI 分析结果列")

        except Exception as e:
            st.error(f"❌ 报告处理失败：{str(e)}")
            import traceback

            st.code(traceback.format_exc())

    elif report_file:
        st.info("👆 请点击上方 **开始 AI 分析** 按钮开始处理")

# ==================== Tab 5: 工具箱 ====================
with tab_tools:
    st.header("🛠️ 常用工具集")
    tool_type = st.radio(
        "选择工具类型",
        ["🧹 Excel 数据清洗", "🎵 音频格式转换 (M4A->WAV)", "📻 音频格式转换 (WAV->PCM)", "💿 音频格式转换 (PCM->WAV)", "📉 音频采样率转换 (48K->16K)", "🕳️ 测试 GAP 分析"],
        horizontal=True
    )

    # --- 1. Excel 数据清洗 (get_dingding_case) ---
    if tool_type == "🧹 Excel 数据清洗":
        st.subheader("🧹 Excel 语料数据清洗")
        st.info("从全功能清单中提取 Case ID、ASR 和中文翻译，支持多语言 (ENU/SPM/PTB/ARG)")
        
        cleaning_file = st.file_uploader("上传 Excel 文件 (.xlsx)", type=["xlsx"], key="cleaning_upload")
        if cleaning_file:
            if st.button("开始清洗", key="start_cleaning_btn"):
                with st.spinner("正在处理..."):
                    try:
                        df = pd.read_excel(cleaning_file, sheet_name='全功能清单', engine='openpyxl')
                        cols_map = {23: 'ENU', 26: 'SPM', 29: 'PTB', 32: 'ARG'}
                        results = {lang: [] for lang in cols_map.values()}
                        error_log = []

                        def is_chinese(text):
                            return len(re.findall(r'[\u4e00-\u9fa5]', str(text))) > 0

                        for index, row in df.iterrows():
                            row_num = index + 2
                            req_id = str(row.iloc[5]).strip() if not pd.isna(row.iloc[5]) else ""
                            for col_idx, lang in cols_map.items():
                                cell_data = row.iloc[col_idx]
                                if pd.isna(cell_data) or str(cell_data).strip() == "": continue
                                lines = str(cell_data).split('\n')
                                for line in lines:
                                    line = line.strip()
                                    if not line or line.lower().startswith('cases:'): continue
                                    first_comma = re.search(r'[,，]', line)
                                    if not first_comma:
                                        error_log.append({'Row': row_num, 'Lang': lang, 'Reason': '缺少分隔符', 'Content': line})
                                        continue
                                    case_id = line[:first_comma.start()].strip()
                                    rest_of_line = line[first_comma.end():].strip()
                                    parts = [p.strip() for p in re.split(r'[,，]', rest_of_line) if p.strip()]
                                    asr_content = ""
                                    zh_translation = ""
                                    found_zh = False
                                    for i, part in enumerate(parts):
                                        if is_chinese(part):
                                            zh_translation = part
                                            asr_content = ", ".join(parts[:i])
                                            found_zh = True
                                            break
                                    if found_zh and asr_content:
                                        results[lang].append({'req_id': req_id, 'case_id': case_id, 'asr': asr_content, 'zh_translation': zh_translation})
                                    else:
                                        reason = "找不到中文翻译" if not found_zh else "ASR内容解析为空"
                                        error_log.append({'Row': row_num, 'Lang': lang, 'Reason': reason, 'Content': line})

                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            for lang, data in results.items():
                                if data: pd.DataFrame(data).to_excel(writer, sheet_name=lang, index=False)
                            if error_log: pd.DataFrame(error_log).to_excel(writer, sheet_name='格式确认', index=False)
                            
                            # 美化
                            workbook = writer.book
                            for sheet in workbook.worksheets:
                                max_col = sheet.max_column
                                max_row = sheet.max_row
                                if max_row > 0: sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
                                for col_cells in sheet.columns:
                                    column_letter = col_cells[0].column_letter
                                    max_width = max(sum(2 if is_chinese(c) else 1 for c in str(cell.value)) for cell in col_cells)
                                    sheet.column_dimensions[column_letter].width = min(max_width + 2, 70)
                        
                        st.success("✅ 清洗完成！")
                        st.download_button("📥 下载结果", data=output.getvalue(), file_name="cleaned_result.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    except Exception as e:
                        st.error(f"❌ 处理失败: {str(e)}")

    # --- 2. M4A to WAV (m4a_to_wav) ---
    elif tool_type == "🎵 音频格式转换 (M4A->WAV)":
        st.subheader("🎵 M4A 转 WAV")
        m4a_files = st.file_uploader("上传 M4A 文件", type=["m4a"], accept_multiple_files=True, key="m4a_upload")
        if m4a_files:
            if st.button("开始转换", key="convert_m4a_btn"):
                progress_bar = st.progress(0)
                converted_files = []
                for i, m4a_file in enumerate(m4a_files):
                    input_path = f"temp_{m4a_file.name}"
                    output_path = os.path.splitext(input_path)[0] + ".wav"
                    with open(input_path, "wb") as f: f.write(m4a_file.getbuffer())
                    
                    command = ['ffmpeg', '-i', input_path, '-acodec', 'pcm_s16le', '-ar', '44100', '-y', output_path]
                    try:
                        subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                        converted_files.append((os.path.splitext(m4a_file.name)[0]+".wav", output_path))
                    except Exception as e:
                        st.error(f"转换 {m4a_file.name} 失败: {e}")
                    finally:
                        if os.path.exists(input_path): os.remove(input_path)
                    progress_bar.progress((i + 1) / len(m4a_files))
                
                if converted_files:
                    st.success(f"✅ 成功转换 {len(converted_files)} 个文件")
                    # 提供打包下载
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for fname, fpath in converted_files:
                            zip_file.write(fpath, fname)
                            os.remove(fpath) # 清理临时文件
                    
                    st.download_button(
                        label="📥 下载全部结果 (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="converted_wav_files.zip",
                        mime="application/zip"
                    )

    # --- 3. WAV to PCM (wav_to_pcm) ---
    elif tool_type == "📻 音频格式转换 (WAV->PCM)":
        st.subheader("📻 WAV 转 PCM (16k/1ch/16bit)")
        wav_files = st.file_uploader("上传 WAV 文件", type=["wav"], accept_multiple_files=True, key="wav_upload")
        if wav_files:
            if st.button("开始转换", key="convert_wav_btn"):
                converted_files = []
                for wav_file in wav_files:
                    wav_path = f"temp_{wav_file.name}"
                    with open(wav_path, "wb") as f: f.write(wav_file.getbuffer())
                    try:
                        with wave.open(wav_path, 'rb') as wf:
                            if wf.getframerate() != 16000 or wf.getnchannels() != 1 or wf.getsampwidth() != 2:
                                st.warning(f"⚠️ {wav_file.name} 不符合 16K/单通道/16bit 要求，已跳过。")
                                continue
                            pcm_data = wf.readframes(wf.getnframes())
                            pcm_filename = os.path.splitext(wav_file.name)[0] + ".pcm"
                            converted_files.append((pcm_filename, pcm_data))
                    except Exception as e:
                        st.error(f"处理 {wav_file.name} 失败: {e}")
                    finally:
                        if os.path.exists(wav_path): os.remove(wav_path)
                
                if converted_files:
                    st.success(f"✅ 成功转换 {len(converted_files)} 个文件")
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for fname, fdata in converted_files:
                            zip_file.writestr(fname, fdata)
                    
                    st.download_button(
                        label="📥 下载全部结果 (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="converted_pcm_files.zip",
                        mime="application/zip"
                    )

    # --- 5. PCM to WAV (AudioConverter_v3 logic) ---
    elif tool_type == "💿 音频格式转换 (PCM->WAV)":
        st.subheader("💿 PCM 转 WAV")
        st.info("将原始 PCM 音频数据封装为标准的 WAV 格式（默认参数：16k/单通道/16bit）")
        
        pcm_files = st.file_uploader("上传 PCM 文件", type=["pcm"], accept_multiple_files=True, key="pcm_upload")
        
        col_pcm1, col_pcm2 = st.columns(2)
        with col_pcm1:
            sample_rate = st.number_input("采样率 (Hz)", value=16000, step=1000)
        with col_pcm2:
            channels = st.number_input("声道数", value=1, min_value=1, max_value=2)
        
        if pcm_files:
            if st.button("开始转换", key="convert_pcm_btn"):
                converted_files = []
                for pcm_file in pcm_files:
                    pcm_path = f"temp_{pcm_file.name}"
                    with open(pcm_path, "wb") as f: f.write(pcm_file.getbuffer())
                    try:
                        wav_filename = os.path.splitext(pcm_file.name)[0] + ".wav"
                        wav_path = f"temp_{wav_filename}"
                        
                        with wave.open(wav_path, 'wb') as wf:
                            wf.setnchannels(channels)
                            wf.setsampwidth(2)  # 16bit
                            wf.setframerate(sample_rate)
                            with open(pcm_path, 'rb') as pf:
                                wf.writeframes(pf.read())
                        
                        converted_files.append((wav_filename, wav_path))
                    except Exception as e:
                        st.error(f"处理 {pcm_file.name} 失败: {e}")
                    finally:
                        if os.path.exists(pcm_path): os.remove(pcm_path)
                
                if converted_files:
                    st.success(f"✅ 成功转换 {len(converted_files)} 个文件")
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for fname, fpath in converted_files:
                            zip_file.write(fpath, fname)
                            os.remove(fpath)
                    
                    st.download_button(
                        label="📥 下载全部结果 (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="converted_wav_from_pcm.zip",
                        mime="application/zip"
                    )

    # --- 6. 48K to 16K (ffmpeg resample) ---
    elif tool_type == "📉 音频采样率转换 (48K->16K)":
        st.subheader("📉 音频采样率转换 (48K -> 16K)")
        st.info("使用 ffmpeg 将高采样率音频重采样为 16kHz，适用于语音识别预处理")
        
        audio_files_48k = st.file_uploader("上传音频文件 (WAV/M4A等)", type=["wav", "m4a", "mp3"], accept_multiple_files=True, key="resample_upload")
        if audio_files_48k:
            if st.button("开始重采样", key="resample_btn"):
                progress_bar = st.progress(0)
                converted_files = []
                for i, audio_file in enumerate(audio_files_48k):
                    input_path = f"temp_{audio_file.name}"
                    output_filename = f"16k_{os.path.splitext(audio_file.name)[0]}.wav"
                    output_path = f"temp_{output_filename}"
                    
                    with open(input_path, "wb") as f: f.write(audio_file.getbuffer())
                    
                    command = ['ffmpeg', '-i', input_path, '-ar', '16000', '-ac', '1', '-y', output_path]
                    try:
                        subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                        converted_files.append((output_filename, output_path))
                    except Exception as e:
                        st.error(f"转换 {audio_file.name} 失败: {e}")
                    finally:
                        if os.path.exists(input_path): os.remove(input_path)
                    progress_bar.progress((i + 1) / len(audio_files_48k))
                
                if converted_files:
                    st.success(f"✅ 成功转换 {len(converted_files)} 个文件")
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for fname, fpath in converted_files:
                            zip_file.write(fpath, fname)
                            os.remove(fpath)
                    
                    st.download_button(
                        label="📥 下载全部结果 (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="resampled_16k_files.zip",
                        mime="application/zip"
                    )

    # --- 7. GAP Analysis (get_gap_id) ---
    elif tool_type == "🕳️ 测试 GAP 分析":
        st.subheader("🕳️ 测试报告 GAP 分析")
        st.info("对比测试报告与优先级清单，找出缺失或 Pass 数量不足的 Req ID")
        
        col_g1, col_g2 = st.columns(2)
        with col_g1:
            report_file_gap = st.file_uploader("上传测试报告 (All Sheet)", type=["xlsx"], key="gap_report")
        with col_g2:
            priority_file_gap = st.file_uploader("上传优先级清单 (ReqID VS P0,1,2,4 Sheet)", type=["xlsx"], key="gap_priority")
        
        priorities = st.multiselect("选择要分析的 Priority", ['P0', 'P1', 'P2', 'P4'], default=['P0', 'P1'])
        min_pass = st.number_input("最小 Pass 数量", min_value=1, value=1)
        
        if report_file_gap and priority_file_gap:
            if st.button("开始分析 GAP", key="analyze_gap_btn"):
                with st.spinner("正在分析..."):
                    try:
                        report_df = pd.read_excel(report_file_gap, sheet_name='All', skiprows=1)
                        priority_df = pd.read_excel(priority_file_gap, sheet_name='ReqID VS P0,1,2,4', skiprows=2)
                        
                        target_req_ids = set()
                        for priority in priorities:
                            filtered_reqs = priority_df[priority_df.iloc[:, 3] == priority].iloc[:, 2]
                            target_req_ids.update(filtered_reqs.dropna().astype(str))
                        
                        req_pass_counts = {}
                        for idx, row in report_df.iterrows():
                            req_id = str(row.iloc[0])
                            final_res = row.iloc[16]
                            if req_id in target_req_ids:
                                if req_id not in req_pass_counts: req_pass_counts[req_id] = 0
                                if final_res == 'P': req_pass_counts[req_id] += 1
                        
                        missing_req_ids = [rid for rid in target_req_ids if rid not in req_pass_counts]
                        insufficient_req_ids = [(rid, cnt) for rid, cnt in req_pass_counts.items() if cnt < min_pass]
                        
                        st.markdown("### 📊 分析结果")
                        if missing_req_ids:
                            st.markdown("**需要补充的 req_id (报告中未找到):**")
                            st.code("\n".join(sorted(missing_req_ids)))
                        else:
                            st.success("所有目标 req_id 在报告中都有对应的 case")
                        
                        if insufficient_req_ids:
                            st.markdown(f"**不满足最小 Pass 数量 ({min_pass}个) 的 req_id:**")
                            df_insuf = pd.DataFrame(insufficient_req_ids, columns=['Req ID', 'Pass Count'])
                            st.dataframe(df_insuf.sort_values(by='Req ID'))
                        else:
                            st.success(f"所有目标 req_id 都满足最小 Pass 数量 ({min_pass}个)")
                    except Exception as e:
                        st.error(f"❌ 分析失败: {str(e)}")

# ==================== 底部信息 ====================

st.markdown("---")
st.caption("Powered by Streamlit + ChromaDB + 千问大模型 | Version 1.0")
